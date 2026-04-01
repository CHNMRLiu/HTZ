"""
HTZ 合同台账管理系统 - 后端 API
v2.0.0 - 全面重构
"""
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, func, case
from typing import Optional, List
from pydantic import BaseModel, Field
import openpyxl
import shutil
import os
import time

from database import init_db, get_db, User, Contract, ContractItem, Invoice
from auth import (
    verify_password, get_password_hash, create_access_token,
    get_current_user, get_admin_user
)

app = FastAPI(title="HTZ 合同台账系统", version="2.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_DIR = "./data/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=UPLOAD_DIR), name="uploads")


@app.on_event("startup")
def startup():
    init_db()
    from database import SessionLocal
    db = SessionLocal()
    try:
        if not db.query(User).first():
            admin = User(
                username="admin",
                hashed_password=get_password_hash("admin123"),
                is_admin=True
            )
            db.add(admin)
            db.commit()
    finally:
        db.close()


# ==================== 工具函数 ====================

def safe_float(v, default=0):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def safe_str(v):
    return str(v).strip() if v is not None and str(v).strip() != "" else None


def api_response(data=None, msg="success", code=200):
    resp = {"code": code, "msg": msg}
    if data is not None:
        resp["data"] = data
    return resp


# ==================== Pydantic 模型 ====================

class LoginRequest(BaseModel):
    username: str
    password: str


class UserCreate(BaseModel):
    username: str
    password: str
    is_admin: bool = False


class UserUpdate(BaseModel):
    password: Optional[str] = None
    is_admin: Optional[bool] = None


class ContractCreate(BaseModel):
    seq: Optional[int] = None
    contract_no: str
    contract_type: str = "采购"
    sign_date: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    amount: Optional[float] = 0
    purchase_amount: Optional[float] = 0
    tax_rate: Optional[float] = 0
    status: str = "进行中"
    buyer: Optional[str] = None
    department: Optional[str] = None
    party_a: Optional[str] = None
    party_b: Optional[str] = None
    content: Optional[str] = None
    invoice_info: Optional[str] = None
    order_progress: Optional[float] = 0
    delivery_progress: Optional[float] = 0


class ContractUpdate(BaseModel):
    seq: Optional[int] = None
    contract_no: Optional[str] = None
    contract_type: Optional[str] = None
    sign_date: Optional[str] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    amount: Optional[float] = None
    purchase_amount: Optional[float] = None
    tax_rate: Optional[float] = None
    status: Optional[str] = None
    buyer: Optional[str] = None
    department: Optional[str] = None
    party_a: Optional[str] = None
    party_b: Optional[str] = None
    content: Optional[str] = None
    invoice_info: Optional[str] = None
    order_progress: Optional[float] = None
    delivery_progress: Optional[float] = None


class BatchStatusUpdate(BaseModel):
    ids: List[int]
    status: str


class ItemCreate(BaseModel):
    item_no: Optional[str] = None
    material_code: Optional[str] = None
    material_name: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = 0
    contract_price: Optional[float] = 0
    total_with_tax: Optional[float] = 0
    purchase_price: Optional[float] = 0
    total_price: Optional[float] = 0
    purchase_contract_no: Optional[str] = None
    order_unit: Optional[str] = None
    sign_date: Optional[str] = None
    delivery_status: Optional[str] = None
    delivery_date: Optional[str] = None
    remark: Optional[str] = None


class InvoiceCreate(BaseModel):
    invoice_no: str
    invoice_date: Optional[str] = None
    amount: Optional[float] = 0
    tax_amount: Optional[float] = 0
    contract_id: Optional[int] = None
    seller: Optional[str] = None


class InvoiceUpdate(BaseModel):
    invoice_no: Optional[str] = None
    invoice_date: Optional[str] = None
    amount: Optional[float] = None
    tax_amount: Optional[float] = None
    contract_id: Optional[int] = None
    seller: Optional[str] = None


# ==================== 认证 ====================

@app.post("/api/auth/login")
def login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.username == req.username).first()
    if not user or not verify_password(req.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="用户名或密码错误")
    token = create_access_token({"sub": user.username})
    return {
        "token": token,
        "username": user.username,
        "is_admin": user.is_admin,
        "id": user.id
    }


@app.get("/api/auth/me")
def get_me(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "username": current_user.username,
        "is_admin": current_user.is_admin
    }


@app.post("/api/auth/users")
def create_user(req: UserCreate, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    if db.query(User).filter(User.username == req.username).first():
        raise HTTPException(status_code=400, detail="用户名已存在")
    user = User(
        username=req.username,
        hashed_password=get_password_hash(req.password),
        is_admin=req.is_admin
    )
    db.add(user)
    db.commit()
    db.refresh(user)
    return api_response({"id": user.id}, msg="创建成功")


@app.get("/api/auth/users")
def list_users(admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    users = db.query(User).order_by(User.id).all()
    return [
        {"id": u.id, "username": u.username, "is_admin": u.is_admin, "created_at": str(u.created_at)}
        for u in users
    ]


@app.put("/api/auth/users/{user_id}")
def update_user(user_id: int, req: UserUpdate, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if user.username == "admin" and req.is_admin is False:
        raise HTTPException(status_code=400, detail="不能取消管理员的管理员权限")
    if req.password:
        user.hashed_password = get_password_hash(req.password)
    if req.is_admin is not None:
        user.is_admin = req.is_admin
    db.commit()
    return api_response(msg="更新成功")


@app.delete("/api/auth/users/{user_id}")
def delete_user(user_id: int, admin: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用户不存在")
    if user.username == "admin":
        raise HTTPException(status_code=400, detail="不能删除默认管理员")
    db.delete(user)
    db.commit()
    return api_response(msg="删除成功")


# ==================== 合同 ====================

def contract_to_dict(c: Contract) -> dict:
    return {
        "id": c.id, "seq": c.seq, "contract_no": c.contract_no,
        "contract_type": c.contract_type or "采购",
        "sign_date": c.sign_date, "start_date": c.start_date, "end_date": c.end_date,
        "amount": c.amount or 0, "purchase_amount": c.purchase_amount or 0,
        "tax_rate": c.tax_rate or 0,
        "status": c.status, "buyer": c.buyer, "department": c.department,
        "party_a": c.party_a, "party_b": c.party_b,
        "content": c.content, "invoice_info": c.invoice_info,
        "order_progress": c.order_progress or 0, "delivery_progress": c.delivery_progress or 0,
        "created_at": str(c.created_at) if c.created_at else None,
        "updated_at": str(c.updated_at) if c.updated_at else None,
    }


@app.get("/api/contracts")
def list_contracts(
    keyword: Optional[str] = None,
    status: Optional[str] = None,
    buyer: Optional[str] = None,
    contract_type: Optional[str] = None,
    department: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = "id",
    sort_order: Optional[str] = "desc",
    page: int = 1,
    size: int = 20,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(Contract)

    # 搜索
    if keyword:
        query = query.filter(or_(
            Contract.contract_no.contains(keyword),
            Contract.content.contains(keyword),
            Contract.buyer.contains(keyword),
            Contract.party_a.contains(keyword),
            Contract.party_b.contains(keyword),
            Contract.department.contains(keyword),
        ))

    # 筛选
    if status:
        query = query.filter(Contract.status == status)
    if buyer:
        query = query.filter(Contract.buyer == buyer)
    if contract_type:
        query = query.filter(Contract.contract_type == contract_type)
    if department:
        query = query.filter(Contract.department == department)
    if date_from:
        query = query.filter(Contract.sign_date >= date_from)
    if date_to:
        query = query.filter(Contract.sign_date <= date_to)

    total = query.count()

    # 排序
    sort_col = getattr(Contract, sort_by, Contract.id)
    if sort_order == "asc":
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    items = query.offset((page - 1) * size).limit(size).all()

    return {
        "total": total,
        "items": [contract_to_dict(c) for c in items]
    }


@app.get("/api/contracts/all")
def list_all_contracts(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取所有合同（用于下拉选择等场景）"""
    contracts = db.query(Contract).order_by(Contract.id.desc()).all()
    return [{"id": c.id, "contract_no": c.contract_no, "buyer": c.buyer} for c in contracts]


@app.get("/api/contracts/stats/monthly")
def monthly_stats(
    year: int = Query(default_factory=lambda: __import__("datetime").datetime.now().year),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """月度合同统计"""
    contracts = db.query(Contract).all()
    monthly = {}
    for m in range(1, 13):
        monthly[m] = {"count": 0, "amount": 0, "purchase_amount": 0}

    for c in contracts:
        if c.sign_date:
            parts = c.sign_date.replace("年", ".").replace("月", ".").replace("日", "").split(".")
            if len(parts) >= 2:
                try:
                    y = int(parts[0])
                    m = int(parts[1])
                    if y == year and 1 <= m <= 12:
                        monthly[m]["count"] += 1
                        monthly[m]["amount"] += (c.amount or 0)
                        monthly[m]["purchase_amount"] += (c.purchase_amount or 0)
                except ValueError:
                    pass

    return [
        {"month": m, **monthly[m]} for m in range(1, 13)
    ]


@app.get("/api/contracts/filters")
def get_filters(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取所有可选的筛选值"""
    buyers = [r[0] for r in db.query(Contract.buyer).distinct().all() if r[0]]
    departments = [r[0] for r in db.query(Contract.department).distinct().all() if r[0]]
    return {"buyers": buyers, "departments": departments}


@app.post("/api/contracts")
def create_contract(data: ContractCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if db.query(Contract).filter(Contract.contract_no == data.contract_no).first():
        raise HTTPException(status_code=400, detail="合同号已存在")
    contract = Contract(**data.model_dump())
    db.add(contract)
    db.commit()
    db.refresh(contract)
    return api_response({"id": contract.id}, msg="创建成功")


@app.get("/api/contracts/{contract_id}")
def get_contract(contract_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    data = contract_to_dict(contract)
    data["items"] = [
        {
            "id": i.id, "item_no": i.item_no, "material_code": i.material_code,
            "material_name": i.material_name, "unit": i.unit, "quantity": i.quantity or 0,
            "contract_price": i.contract_price or 0, "total_with_tax": i.total_with_tax or 0,
            "purchase_price": i.purchase_price or 0, "total_price": i.total_price or 0,
            "purchase_contract_no": i.purchase_contract_no, "order_unit": i.order_unit,
            "sign_date": i.sign_date, "delivery_status": i.delivery_status,
            "delivery_date": i.delivery_date, "remark": i.remark,
        }
        for i in contract.items
    ]
    return data


@app.put("/api/contracts/{contract_id}")
def update_contract(contract_id: int, data: ContractUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    update_data = data.model_dump(exclude_unset=True)
    if "contract_no" in update_data:
        existing = db.query(Contract).filter(
            Contract.contract_no == update_data["contract_no"],
            Contract.id != contract_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="合同号已存在")
    for key, value in update_data.items():
        setattr(contract, key, value)
    db.commit()
    return api_response(msg="更新成功")


@app.delete("/api/contracts/{contract_id}")
def delete_contract(contract_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    db.delete(contract)
    db.commit()
    return api_response(msg="删除成功")


@app.post("/api/contracts/batch-delete")
def batch_delete_contracts(data: BatchStatusUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    count = db.query(Contract).filter(Contract.id.in_(data.ids)).delete(synchronize_session=False)
    db.commit()
    return api_response(msg=f"已删除 {count} 条", data={"count": count})


@app.post("/api/contracts/batch-status")
def batch_update_status(data: BatchStatusUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    count = db.query(Contract).filter(Contract.id.in_(data.ids)).update(
        {"status": data.status}, synchronize_session=False
    )
    db.commit()
    return api_response(msg=f"已更新 {count} 条", data={"count": count})


# ==================== 合同明细 ====================

@app.post("/api/contracts/{contract_id}/items")
def create_item(contract_id: int, data: ItemCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")
    item = ContractItem(contract_id=contract_id, **data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return api_response({"id": item.id}, msg="添加成功")


@app.put("/api/items/{item_id}")
def update_item(item_id: int, data: ItemCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    item = db.query(ContractItem).filter(ContractItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="明细不存在")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(item, key, value)
    db.commit()
    return api_response(msg="更新成功")


@app.delete("/api/items/{item_id}")
def delete_item(item_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    item = db.query(ContractItem).filter(ContractItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="明细不存在")
    db.delete(item)
    db.commit()
    return api_response(msg="删除成功")


@app.post("/api/contracts/{contract_id}/items/import")
async def import_items(contract_id: int, file: UploadFile = File(...), current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    filepath = os.path.join(UPLOAD_DIR, f"import_items_{contract_id}_{int(time.time())}.xlsx")
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    count = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            item = ContractItem(
                contract_id=contract_id,
                item_no=safe_str(row[0]),
                material_code=safe_str(row[1]),
                material_name=safe_str(row[2]),
                unit=safe_str(row[3]),
                quantity=safe_float(row[4]),
                contract_price=safe_float(row[5]),
                total_with_tax=safe_float(row[6]),
                purchase_price=safe_float(row[7]),
                total_price=safe_float(row[8]),
                purchase_contract_no=safe_str(row[9]) if len(row) > 9 else None,
                order_unit=safe_str(row[10]) if len(row) > 10 else None,
                sign_date=safe_str(row[11]) if len(row) > 11 else None,
                delivery_status=safe_str(row[12]) if len(row) > 12 else None,
                delivery_date=safe_str(row[13]) if len(row) > 13 else None,
                remark=safe_str(row[14]) if len(row) > 14 else None,
            )
            db.add(item)
            count += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    db.commit()

    # 清理临时文件
    try:
        os.remove(filepath)
    except:
        pass

    return api_response({"imported": count, "errors": errors}, msg=f"导入完成，成功 {count} 条")


@app.get("/api/contracts/{contract_id}/items/export")
def export_items(contract_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    contract = db.query(Contract).filter(Contract.id == contract_id).first()
    if not contract:
        raise HTTPException(status_code=404, detail="合同不存在")

    items = db.query(ContractItem).filter(ContractItem.contract_id == contract_id).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同明细"

    headers = ["项目号", "物料编码", "物料名称", "单位", "数量", "合同单价", "总计（含税）", "采购价", "总价", "采购合同号", "订货单位", "签订日期", "交货情况", "交货期", "备注"]
    ws.append(headers)

    for i in items:
        ws.append([i.item_no, i.material_code, i.material_name, i.unit, i.quantity, i.contract_price, i.total_with_tax, i.purchase_price, i.total_price, i.purchase_contract_no, i.order_unit, i.sign_date, i.delivery_status, i.delivery_date, i.remark])

    total_sales = sum((i.total_with_tax or 0) for i in items)
    total_purchase = sum((i.total_price or 0) for i in items)
    ws.append([])
    ws.append(["", "", "", "", "", "", "销售金额合计:", "", total_sales])
    ws.append(["", "", "", "", "", "", "采购金额合计:", "", total_purchase])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16

    filepath = os.path.join(UPLOAD_DIR, f"export_items_{contract_id}.xlsx")
    wb.save(filepath)

    return FileResponse(filepath, filename=f"{contract.contract_no}_明细.xlsx")


# ==================== 模板下载 ====================

@app.get("/api/template/items")
def download_item_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同明细模板"

    headers = ["项目号", "物料编码", "物料名称", "单位", "数量", "合同单价", "总计（含税）", "采购价", "总价", "采购合同号", "订货单位", "签订日期", "交货情况", "交货期", "备注"]
    ws.append(headers)
    ws.append(["1", "MAT001", "示例物料名称", "台", 2, 5000, 10000, 4500, 9000, "PO2025001", "示例公司", "2025.1.1", "未交货", "30天", ""])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16

    filepath = os.path.join(UPLOAD_DIR, "template_items.xlsx")
    wb.save(filepath)
    return FileResponse(filepath, filename="合同明细导入模板.xlsx")


@app.get("/api/template/contracts")
def download_contract_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同导入模板"

    headers = ["序号", "供应合同号", "合同类型", "签订日期", "履约开始日", "到期日期", "合同金额", "采购金额", "税率%", "合同状态", "采购员", "归属部门", "甲方", "乙方", "合同内容", "开票信息"]
    ws.append(headers)
    ws.append([1, "2501CBHW0001", "采购", "2025.1.1", "2025.1.1", "2025.12.31", 10000, 8000, 13, "进行中", "文旭", "生产部", "长沙水泵厂", "供应商", "示例合同内容", ""])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    filepath = os.path.join(UPLOAD_DIR, "template_contracts.xlsx")
    wb.save(filepath)
    return FileResponse(filepath, filename="合同导入模板.xlsx")


# ==================== 发票 ====================

def invoice_to_dict(i: Invoice) -> dict:
    return {
        "id": i.id, "invoice_no": i.invoice_no, "invoice_date": i.invoice_date,
        "amount": i.amount or 0, "tax_amount": i.tax_amount or 0,
        "contract_id": i.contract_id, "seller": i.seller,
        "file_path": i.file_path,
        "attachment_path": i.attachment_path, "attachment_name": i.attachment_name,
        "contract_no": i.contract.contract_no if i.contract else None,
        "created_at": str(i.created_at) if i.created_at else None,
    }


@app.get("/api/invoices")
def list_invoices(
    keyword: Optional[str] = None,
    contract_id: Optional[int] = None,
    seller: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort_by: Optional[str] = "id",
    sort_order: Optional[str] = "desc",
    page: int = 1,
    size: int = 20,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    query = db.query(Invoice)

    if keyword:
        query = query.filter(or_(
            Invoice.invoice_no.contains(keyword),
            Invoice.seller.contains(keyword),
        ))
    if contract_id:
        query = query.filter(Invoice.contract_id == contract_id)
    if seller:
        query = query.filter(Invoice.seller == seller)
    if date_from:
        query = query.filter(Invoice.invoice_date >= date_from)
    if date_to:
        query = query.filter(Invoice.invoice_date <= date_to)

    total = query.count()

    sort_col = getattr(Invoice, sort_by, Invoice.id)
    if sort_order == "asc":
        query = query.order_by(sort_col.asc())
    else:
        query = query.order_by(sort_col.desc())

    items = query.offset((page - 1) * size).limit(size).all()

    return {
        "total": total,
        "items": [invoice_to_dict(i) for i in items]
    }


@app.post("/api/invoices")
def create_invoice(data: InvoiceCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    if db.query(Invoice).filter(Invoice.invoice_no == data.invoice_no).first():
        raise HTTPException(status_code=400, detail="发票号已存在")
    invoice = Invoice(**data.model_dump())
    db.add(invoice)
    db.commit()
    db.refresh(invoice)
    return api_response({"id": invoice.id}, msg="创建成功")


@app.get("/api/invoices/{invoice_id}")
def get_invoice(invoice_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    return invoice_to_dict(invoice)


@app.put("/api/invoices/{invoice_id}")
def update_invoice(invoice_id: int, data: InvoiceUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    update_data = data.model_dump(exclude_unset=True)
    if "invoice_no" in update_data:
        existing = db.query(Invoice).filter(
            Invoice.invoice_no == update_data["invoice_no"],
            Invoice.id != invoice_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="发票号已存在")
    for key, value in update_data.items():
        setattr(invoice, key, value)
    db.commit()
    return api_response(msg="更新成功")


@app.delete("/api/invoices/{invoice_id}")
def delete_invoice(invoice_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="发票不存在")
    db.delete(invoice)
    db.commit()
    return api_response(msg="删除成功")


@app.post("/api/invoices/upload")
async def upload_invoice_attachment(
    file: UploadFile = File(...),
    invoice_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    attach_dir = os.path.join(UPLOAD_DIR, "invoice_attachments")
    os.makedirs(attach_dir, exist_ok=True)

    original_name = file.filename or "attachment"
    if invoice_id:
        invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if invoice:
            safe_name = f"{invoice.invoice_no}_{original_name}"
            filepath = os.path.join(attach_dir, safe_name)
            with open(filepath, "wb") as f:
                shutil.copyfileobj(file.file, f)
            invoice.attachment_path = f"/uploads/invoice_attachments/{safe_name}"
            invoice.attachment_name = original_name
            db.commit()
            return api_response({"path": invoice.attachment_path, "name": original_name}, msg="上传成功")

    safe_name = f"{int(time.time())}_{original_name}"
    filepath = os.path.join(attach_dir, safe_name)
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)
    return api_response({"path": f"/uploads/invoice_attachments/{safe_name}", "name": original_name}, msg="上传成功")


# ==================== 导入导出 ====================

@app.post("/api/import/contracts")
async def import_contracts(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    filepath = os.path.join(UPLOAD_DIR, f"import_contracts_{int(time.time())}.xlsx")
    with open(filepath, "wb") as f:
        shutil.copyfileobj(file.file, f)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    count = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            def get_col(idx, default=None):
                return row[idx] if len(row) > idx and row[idx] is not None else default

            seq = get_col(0)
            contract_no = str(get_col(1, "")).strip()
            if not contract_no:
                errors.append(f"第{row_idx}行: 合同号为空")
                continue

            contract_type = str(get_col(2, "采购")).strip()
            sign_date = safe_str(get_col(3))
            start_date = safe_str(get_col(4))
            end_date = safe_str(get_col(5))
            amount = safe_float(get_col(6))
            purchase_amount = safe_float(get_col(7))
            tax_rate = safe_float(get_col(8))
            status_val = str(get_col(9, "进行中")).strip()
            buyer = safe_str(get_col(10))
            department = safe_str(get_col(11))
            party_a = safe_str(get_col(12))
            party_b = safe_str(get_col(13))
            content = safe_str(get_col(14))
            invoice_info = safe_str(get_col(15))

            # 兼容旧状态映射
            status_map = {"已开票": "待回款", "已交货": "待开票", "未交货": "进行中"}
            status_val = status_map.get(status_val, status_val)

            existing = db.query(Contract).filter(Contract.contract_no == contract_no).first()
            if existing:
                existing.seq = seq
                existing.contract_type = contract_type
                existing.sign_date = sign_date
                existing.start_date = start_date
                existing.end_date = end_date
                existing.amount = amount
                existing.purchase_amount = purchase_amount
                existing.tax_rate = tax_rate
                existing.status = status_val
                existing.buyer = buyer
                existing.department = department
                existing.party_a = party_a
                existing.party_b = party_b
                existing.content = content
                existing.invoice_info = invoice_info
            else:
                contract = Contract(
                    seq=seq, contract_no=contract_no, contract_type=contract_type,
                    sign_date=sign_date, start_date=start_date, end_date=end_date,
                    amount=amount, purchase_amount=purchase_amount, tax_rate=tax_rate,
                    status=status_val, buyer=buyer, department=department,
                    party_a=party_a, party_b=party_b, content=content, invoice_info=invoice_info
                )
                db.add(contract)
            count += 1
        except Exception as e:
            errors.append(f"第{row_idx}行: {str(e)}")

    db.commit()

    try:
        os.remove(filepath)
    except:
        pass

    return api_response({"imported": count, "errors": errors}, msg=f"导入完成，成功 {count} 条")


@app.get("/api/export/contracts")
def export_contracts(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    contracts = db.query(Contract).order_by(Contract.seq.asc().nullslast(), Contract.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同台账"

    headers = ["序号", "供应合同号", "合同类型", "签订日期", "履约开始日", "到期日期", "合同金额", "采购金额", "税率%", "合同状态", "采购员", "归属部门", "甲方", "乙方", "合同内容", "开票信息", "订货进度%", "交货进度%"]
    ws.append(headers)

    for c in contracts:
        ws.append([c.seq, c.contract_no, c.contract_type or "采购", c.sign_date, c.start_date, c.end_date,
                    c.amount or 0, c.purchase_amount or 0, c.tax_rate or 0, c.status, c.buyer, c.department,
                    c.party_a, c.party_b, c.content, c.invoice_info, c.order_progress or 0, c.delivery_progress or 0])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16

    filepath = os.path.join(UPLOAD_DIR, "export_contracts.xlsx")
    wb.save(filepath)
    return FileResponse(filepath, filename="合同台账.xlsx")


@app.get("/api/export/invoices")
def export_invoices(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    invoices = db.query(Invoice).order_by(Invoice.id.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "发票台账"

    headers = ["发票号码", "开票日期", "价税合计", "税额", "关联合同号", "销售方"]
    ws.append(headers)

    for i in invoices:
        contract_no = i.contract.contract_no if i.contract else ""
        ws.append([i.invoice_no, i.invoice_date, i.amount or 0, i.tax_amount or 0, contract_no, i.seller])

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18

    filepath = os.path.join(UPLOAD_DIR, "export_invoices.xlsx")
    wb.save(filepath)
    return FileResponse(filepath, filename="发票台账.xlsx")


# ==================== 统计 ====================

@app.get("/api/stats")
def get_stats(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    total_contracts = db.query(Contract).count()
    total_invoices = db.query(Invoice).count()

    total_amount = db.query(func.sum(Contract.amount)).scalar() or 0
    total_purchase = db.query(func.sum(Contract.purchase_amount)).scalar() or 0
    total_invoice_amount = db.query(func.sum(Invoice.amount)).scalar() or 0
    total_invoice_tax = db.query(func.sum(Invoice.tax_amount)).scalar() or 0

    # 按状态统计
    by_status = {}
    status_rows = db.query(Contract.status, func.count(Contract.id)).group_by(Contract.status).all()
    for status, count in status_rows:
        by_status[status or "未知"] = count

    # 按采购员统计
    by_buyer = {}
    buyer_rows = db.query(
        Contract.buyer,
        func.count(Contract.id),
        func.sum(Contract.amount),
        func.sum(Contract.purchase_amount)
    ).group_by(Contract.buyer).all()
    for buyer, count, amount, purchase_amount in buyer_rows:
        by_buyer[buyer or "未知"] = {
            "count": count,
            "amount": amount or 0,
            "purchase_amount": purchase_amount or 0,
        }

    # 按合同类型统计
    by_type = {}
    type_rows = db.query(Contract.contract_type, func.count(Contract.id)).group_by(Contract.contract_type).all()
    for ctype, count in type_rows:
        by_type[ctype or "其他"] = count

    # 到期合同（未来30天内）
    from datetime import datetime, timedelta
    today = datetime.now().strftime("%Y.%m.%d")
    thirty_days = (datetime.now() + timedelta(days=30)).strftime("%Y.%m.%d")
    expiring_soon = db.query(Contract).filter(
        Contract.end_date.isnot(None),
        Contract.end_date > today,
        Contract.end_date <= thirty_days,
        Contract.status != "已完结"
    ).count()

    expired = db.query(Contract).filter(
        Contract.end_date.isnot(None),
        Contract.end_date <= today,
        Contract.status != "已完结"
    ).count()

    # 明细汇总
    total_items = db.query(ContractItem).count()
    total_item_sales = db.query(func.sum(ContractItem.total_with_tax)).scalar() or 0
    total_item_purchase = db.query(func.sum(ContractItem.total_price)).scalar() or 0

    return {
        "total_contracts": total_contracts,
        "total_invoices": total_invoices,
        "total_amount": total_amount,
        "total_purchase": total_purchase,
        "total_invoice_amount": total_invoice_amount,
        "total_invoice_tax": total_invoice_tax,
        "total_profit": total_amount - total_purchase,
        "by_status": by_status,
        "by_buyer": by_buyer,
        "by_type": by_type,
        "expiring_soon": expiring_soon,
        "expired": expired,
        "total_items": total_items,
        "total_item_sales": total_item_sales,
        "total_item_purchase": total_item_purchase,
    }
